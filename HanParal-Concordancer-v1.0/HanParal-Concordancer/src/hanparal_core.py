"""
HanParal — Multilingual Literary Concordancer
Core functions — v1.0

A Python-based concordance tool for comparative translation analysis
of culture-specific items (CSIs) in manually aligned literary corpora.

Author : Lívia Aguiar Correia Cavalcanti
ORCID  : https://orcid.org/0000-0003-3707-9966
License: MIT © 2026
"""

import re
import unicodedata
from pathlib import Path
from typing import Dict, Iterable

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

import matplotlib.colors as mcolors
from matplotlib.colors import LinearSegmentedColormap

HANPARAL_VERSION = "1.0.0"

# ── HanParal palette — Sienna Tijolo ─────────────────────────
HP_COLORS     = {"target_1":"#8B4A38","target_2":"#C07A5A","target_3":"#E0B090"}
HP_BG         = "#FFF9F7"
HP_SPINE      = "#E8C8BC"
HP_GRID       = "#F0DDD7"
HP_TEXT       = "#3A1A10"
HP_TEXT_MUTED = "#8B6A60"
HP_EMPTY      = "#F5F0EE"

HP_CMAP = LinearSegmentedColormap.from_list(
    "hanparal_sienna",
    ["#FFF9F7", "#F0DDD7", "#C07A5A", "#8B4A38"],
    N=256,
)


# ============================================================
# 1. Column schema / target-text configuration
# ============================================================

# Change these numbers according to your project.
# Example:
# N_TARGETS = 1  -> target_1
# N_TARGETS = 2  -> target_1, target_2
# N_TARGETS = 3  -> target_1, target_2, target_3
#
# N_STRATEGY_SLOTS controls how many dropdown slots each target text receives.
# Example with N_TARGETS = 3 and N_STRATEGY_SLOTS = 3:
# strategy_target_1_1, strategy_target_1_2, strategy_target_1_3
# strategy_target_2_1, strategy_target_2_2, strategy_target_2_3
# strategy_target_3_1, strategy_target_3_2, strategy_target_3_3
N_TARGETS = 3
N_STRATEGY_SLOTS = 3


SOURCE_COLUMNS = [
    "id",
    "section",
    "source_text",
]


HP_TARGET_PALETTE = [
    "#8B4A38",
    "#C07A5A",
    "#E0B090",
    "#A05C47",
    "#D7A184",
    "#6F3A2D",
]


def configure_target_columns(n_targets=3, n_strategy_slots=3):
    """
    Configure HanParal according to the number of target texts and
    the number of manual strategy slots per target.

    Minimal required input after configuration:
    id | section | source_text | target_1 | target_2 | ...

    Annotation output:
    strategy_target_1_1 | strategy_target_1_2 | ...
    strategy_target_2_1 | strategy_target_2_2 | ...
    """
    if not isinstance(n_targets, int) or n_targets < 1:
        raise ValueError("n_targets must be an integer greater than or equal to 1.")

    if not isinstance(n_strategy_slots, int) or n_strategy_slots < 1:
        raise ValueError("n_strategy_slots must be an integer greater than or equal to 1.")

    global N_TARGETS
    global N_STRATEGY_SLOTS
    global TARGET_COLUMNS
    global STRATEGY_COLUMNS
    global STRATEGY_MAP
    global LANGUAGE_COLUMNS
    global BASE_COLUMNS
    global HP_COLORS

    N_TARGETS = n_targets
    N_STRATEGY_SLOTS = n_strategy_slots

    TARGET_COLUMNS = [
        f"target_{i}"
        for i in range(1, N_TARGETS + 1)
    ]

    STRATEGY_COLUMNS = [
        f"strategy_target_{target_i}_{slot_i}"
        for target_i in range(1, N_TARGETS + 1)
        for slot_i in range(1, N_STRATEGY_SLOTS + 1)
    ]

    STRATEGY_MAP = {
        target_col: [
            f"strategy_target_{target_i}_{slot_i}"
            for slot_i in range(1, N_STRATEGY_SLOTS + 1)
        ]
        for target_i, target_col in enumerate(TARGET_COLUMNS, start=1)
    }

    LANGUAGE_COLUMNS = {
        target_col: target_col
        for target_col in TARGET_COLUMNS
    }

    BASE_COLUMNS = SOURCE_COLUMNS + TARGET_COLUMNS

    HP_COLORS = {
        target_col: HP_TARGET_PALETTE[(i - 1) % len(HP_TARGET_PALETTE)]
        for i, target_col in enumerate(TARGET_COLUMNS, start=1)
    }


configure_target_columns(N_TARGETS, N_STRATEGY_SLOTS)



def parse_items_text(text):
    """
    Convert a comma/semicolon/newline-separated string into a clean list.

    Example:
        "누나, 형, 도청" -> ["누나", "형", "도청"]
    """
    if text is None:
        return []

    parts = re.split(r"[,;\n]+", str(text))

    return [
        part.strip()
        for part in parts
        if part.strip()
    ]


def make_safe_filename_fragment(items, default="analysis", max_length=80):
    """
    Create a safe filename fragment from one item or a list of items.
    """
    if isinstance(items, str):
        raw = items
    else:
        raw = "_".join(str(item).strip() for item in items if str(item).strip())

    raw = raw.strip() or default
    safe = re.sub(r"[^\w가-힣一-龥ぁ-んァ-ヶ]+", "_", raw, flags=re.UNICODE)
    safe = re.sub(r"_+", "_", safe).strip("_")

    return (safe or default)[:max_length]


# ============================================================
# 2. Validation and loading
# ============================================================

def validate_base_columns(df):
    """
    Validate the minimal HanParal input schema.

    Required columns:
    - id
    - section
    - source_text
    - target_1, target_2, ...

    Optional columns such as term_source and category are not required
    in the base corpus. They are generated or preserved later in Phase 1.
    """
    missing = [
        col for col in BASE_COLUMNS
        if col not in df.columns
    ]

    if missing:
        raise ValueError(
            "Missing required column(s): "
            + ", ".join(missing)
            + "\n\nRequired input columns are: "
            + ", ".join(BASE_COLUMNS)
        )

    return True

def load_corpus(path, sheet_name=0):
    """
    Load an aligned corpus from Excel or CSV.

    Accepted formats:
    - .xlsx
    - .xls
    - .xlsm
    - .csv

    Required columns:
    id, section, source_text, target_1, target_2, ...

    Optional columns:
    term_source, category, strategy_target_1_1, strategy_target_1_2, ...
    """
    path = str(path)
    path_lower = path.lower()

    if path_lower.endswith((".xlsx", ".xls", ".xlsm")):
        df = pd.read_excel(path, sheet_name=sheet_name)
    elif path_lower.endswith(".csv"):
        df = pd.read_csv(path)
    else:
        raise ValueError("Use .xlsx, .xls, .xlsm, or .csv")

    df.columns = [str(c).strip() for c in df.columns]

    validate_base_columns(df)

    return df


# ============================================================
# 3. Text normalization
# ============================================================

def normalize_text(text, remove_accents=True, lowercase=True):
    """
    Normalize text for easier equivalent detection.

    Useful for Portuguese/Spanish forms with accents.
    """
    text = "" if pd.isna(text) else str(text)

    if lowercase:
        text = text.lower()

    if remove_accents:
        text = "".join(
            ch for ch in unicodedata.normalize("NFD", text)
            if unicodedata.category(ch) != "Mn"
        )

    return text


def contains_form(text, form, exact_word=False, remove_accents=True):
    """
    Check whether a translation contains an expected equivalent form.
    """
    text_norm = normalize_text(text, remove_accents=remove_accents)
    form_norm = normalize_text(form, remove_accents=remove_accents)

    if not form_norm:
        return False

    if exact_word:
        return bool(
            re.search(r"\b" + re.escape(form_norm) + r"\b", text_norm)
        )

    return form_norm in text_norm


# ============================================================
# 4. Occurrence search
# ============================================================

def _ordered_unique_columns(columns):
    """
    Return column names without duplicates while preserving order.
    """
    seen = set()
    out = []

    for col in columns:
        if col not in seen:
            seen.add(col)
            out.append(col)

    return out


def _add_phase1_metadata_columns(df, term):
    """
    Add Phase 1 metadata columns to search results.

    The base corpus does not need term_source or category.
    After searching, HanParal creates or updates:

    term_source | category | ...

    - term_source receives the term searched at runtime.
    - category is created empty for the researcher to fill manually.
    """
    out = df.copy()

    term = str(term).strip()

    # term_source is generated from the searched term.
    if "term_source" in out.columns:
        out["term_source"] = term
    else:
        out.insert(0, "term_source", term)

    # category is created empty for researcher annotation.
    # If a pre-tagged category column already exists, it is preserved.
    if "category" not in out.columns:
        insert_position = out.columns.get_loc("term_source") + 1
        out.insert(insert_position, "category", "")

    first_cols = ["term_source", "category"]
    other_cols = [
        col for col in out.columns
        if col not in first_cols
    ]

    out = out[_ordered_unique_columns(first_cols + other_cols)]

    return out



def search_source_term(
    df,
    term,
    search_in=("source_text",),
    case_sensitive=True,
    exact_term_field=False
):
    """
    Search one source-language term in the aligned corpus.

    The searched term is NOT required in the input file.
    It is supplied at runtime and added automatically to the output
    as term_source.

    Default search column:
    - source_text

    The input corpus can be as simple as:
    id | section | source_text | target_1 | target_2 | ...

    Adds:
    - source_row_number: original spreadsheet-like row number
      from the input corpus, used for safe alignment collapse.
    """
    term = str(term).strip()

    if not term:
        raise ValueError("Please provide a source term to search.")

    search_columns = [
        col for col in search_in
        if col in df.columns
    ]

    if not search_columns:
        raise ValueError(
            "None of the requested search columns exist in the corpus: "
            + ", ".join(search_in)
            + "\n\nFor minimal input, use search_in=(\"source_text\",)."
        )

    mask = pd.Series(False, index=df.index)

    for col in search_columns:
        series = df[col].fillna("").astype(str)

        if col == "term_source" and exact_term_field:
            if case_sensitive:
                mask = mask | (series.str.strip() == term)
            else:
                mask = mask | (series.str.strip().str.lower() == term.lower())
        else:
            mask = mask | series.str.contains(
                re.escape(term),
                na=False,
                regex=True,
                case=case_sensitive
            )

    results = df.loc[mask].copy()

    output_columns = _ordered_unique_columns(
        ["term_source", "category", "source_row_number"] + list(df.columns)
    )

    if results.empty:
        return pd.DataFrame(columns=output_columns)

    # Store input row number for traceability and safe alignment collapse.
    # +2 approximates the spreadsheet row number when row 1 is the header.
    if "source_row_number" not in results.columns:
        results.insert(0, "source_row_number", results.index.astype(int) + 2)

    results = results.reset_index(drop=True)

    results = _add_phase1_metadata_columns(results, term)

    return results

def search_multiple_source_terms(
    df,
    terms,
    search_in=("source_text",),
    case_sensitive=True,
    exact_term_field=False
):
    """
    Search multiple source-language terms at once.

    The searched terms are NOT required in the input file.
    Each result row receives a generated term_source value.

    If the same corpus row matches more than one searched term,
    it appears once for each matched term. This is intentional for annotation.
    """
    cleaned_terms = [
        str(term).strip()
        for term in terms
        if str(term).strip()
    ]

    if not cleaned_terms:
        raise ValueError("Please provide at least one source term.")

    all_results = []

    for term in cleaned_terms:
        term_results = search_source_term(
            df,
            term,
            search_in=search_in,
            case_sensitive=case_sensitive,
            exact_term_field=exact_term_field
        )

        if not term_results.empty:
            all_results.append(term_results)

    output_columns = _ordered_unique_columns(
        ["term_source", "category"] + list(df.columns)
    )

    if not all_results:
        return pd.DataFrame(columns=output_columns)

    combined = pd.concat(all_results, ignore_index=True)

    return combined

def count_term_hits_in_text(
    text,
    term,
    case_sensitive=True
):
    """
    Count how many times a searched term appears inside one source_text cell.

    This counts token hits, not just matching rows/passages.
    Example:
    source_text = "언니가 말했다. 언니는 웃었다."
    term = "언니"
    result = 2
    """
    if pd.isna(text) or pd.isna(term):
        return 0

    text = str(text)
    term = str(term).strip()

    if not term:
        return 0

    flags = 0 if case_sensitive else re.IGNORECASE

    return len(
        re.findall(
            re.escape(term),
            text,
            flags=flags
        )
    )


def add_token_hit_counts(
    df,
    case_sensitive=True
):
    """
    Add token_hit_count to search results.

    token_hit_count = how many times term_source appears inside source_text.
    """
    out = df.copy()

    if out.empty:
        out["token_hit_count"] = []
        return out

    if "term_source" not in out.columns:
        raise ValueError("Column 'term_source' not found.")

    if "source_text" not in out.columns:
        raise ValueError("Column 'source_text' not found.")

    out["token_hit_count"] = out.apply(
        lambda row: count_term_hits_in_text(
            row["source_text"],
            row["term_source"],
            case_sensitive=case_sensitive
        ),
        axis=1
    )

    return out



def expand_search_results_to_token_hits(
    df,
    context_chars=35,
    case_sensitive=True
):
    """
    Expand passage-level search results into token-level rows.

    One output row = one token hit.

    This is the recommended HanParal Phase 1 output for literary
    translation analysis because different occurrences of the same term
    inside one source segment may receive different translation strategies.
    """
    if df.empty:
        out = df.copy()
        expected_cols = (
            ["term_source", "category", "token_hit_id", "token_hit_index",
             "token_hit_count_in_segment", "token_start", "token_end",
             "kwic_source", "id", "section", "source_text"]
            + TARGET_COLUMNS
        )
        return pd.DataFrame(columns=_ordered_unique_columns(expected_cols + list(out.columns)))

    if "term_source" not in df.columns:
        raise ValueError("Column 'term_source' not found.")

    if "source_text" not in df.columns:
        raise ValueError("Column 'source_text' not found.")

    rows = []
    flags = 0 if case_sensitive else re.IGNORECASE

    for _, row in df.iterrows():
        term = "" if pd.isna(row.get("term_source", "")) else str(row.get("term_source", "")).strip()
        text = "" if pd.isna(row.get("source_text", "")) else str(row.get("source_text", ""))

        if not term:
            continue

        matches = list(
            re.finditer(
                re.escape(term),
                text,
                flags=flags
            )
        )

        if not matches:
            continue

        total_hits = len(matches)

        for hit_index, match in enumerate(matches, start=1):
            start = max(0, match.start() - context_chars)
            end = min(len(text), match.end() + context_chars)

            kwic = (
                ("…" if start > 0 else "")
                + text[start:end]
                + ("…" if end < len(text) else "")
            )

            new_row = row.copy()

            if "category" not in new_row.index:
                new_row["category"] = ""

            base_id = str(row.get("source_occurrence_id", row.get("id", ""))).strip()
            safe_term = make_safe_filename_fragment(term, default="term")

            new_row["token_hit_id"] = f"{base_id}__{safe_term}__hit_{hit_index:02d}"
            new_row["token_hit_index"] = hit_index
            new_row["token_hit_count_in_segment"] = total_hits
            new_row["token_start"] = match.start()
            new_row["token_end"] = match.end()
            new_row["kwic_source"] = kwic

            rows.append(new_row)

    if not rows:
        expected_cols = (
            ["term_source", "category", "token_hit_id", "token_hit_index",
             "token_hit_count_in_segment", "token_start", "token_end",
             "kwic_source", "id", "section", "source_text"]
            + TARGET_COLUMNS
        )
        return pd.DataFrame(columns=expected_cols)

    out = pd.DataFrame(rows).reset_index(drop=True)

    preferred_order = (
        ["term_source", "category"]
        + (["source_occurrence_id"] if "source_occurrence_id" in out.columns else [])
        + (["alignment_rows_collapsed"] if "alignment_rows_collapsed" in out.columns else [])
        + [
            "token_hit_id",
            "token_hit_index",
            "token_hit_count_in_segment",
            "token_start",
            "token_end",
            "kwic_source",
        ]
        + (["source_row_number"] if "source_row_number" in out.columns else [])
        + ["id", "section", "source_text"]
        + TARGET_COLUMNS
    )

    existing_preferred_order = [
        col for col in preferred_order
        if col in out.columns
    ]

    remaining_cols = [
        col for col in out.columns
        if col not in existing_preferred_order
    ]

    out = out[existing_preferred_order + remaining_cols]

    return out

def make_kwic_by_term_source(df, context_chars=35):
    """
    Create KWIC context using each row's term_source.

    This is useful after single-term search and multi-term search,
    because term_source is generated from the term supplied at runtime.
    """
    if "term_source" not in df.columns:
        raise ValueError("Column 'term_source' not found.")

    if "source_text" not in df.columns:
        raise ValueError("Column 'source_text' not found.")

    def extract(row):
        term = "" if pd.isna(row["term_source"]) else str(row["term_source"]).strip()
        text = "" if pd.isna(row["source_text"]) else str(row["source_text"])

        if not term:
            return text[:context_chars * 2]

        pos = text.find(term)

        if pos == -1:
            return text[:context_chars * 2]

        start = max(0, pos - context_chars)
        end = min(len(text), pos + len(term) + context_chars)

        return (
            ("…" if start > 0 else "")
            + text[start:end]
            + ("…" if end < len(text) else "")
        )

    out = df.copy()

    kwic_values = out.apply(extract, axis=1)

    if "kwic_source" in out.columns:
        out = out.drop(columns=["kwic_source"])

    if "category" in out.columns:
        insert_position = out.columns.get_loc("category") + 1
    elif "term_source" in out.columns:
        insert_position = out.columns.get_loc("term_source") + 1
    else:
        insert_position = 0

    out.insert(insert_position, "kwic_source", kwic_values)

    return out


def make_kwic_by_searched_term(df, context_chars=35):
    """
    Backward-compatible alias for older notebooks.

    New HanParal outputs use term_source rather than searched_term.
    """
    if "term_source" in df.columns:
        return make_kwic_by_term_source(df, context_chars=context_chars)

    if "searched_term" not in df.columns:
        raise ValueError("Column 'term_source' or 'searched_term' not found.")

    out = df.copy()
    out = out.rename(columns={"searched_term": "term_source"})

    return make_kwic_by_term_source(out, context_chars=context_chars)


def make_kwic(df, term=None, context_chars=35):
    """
    Create a KWIC-like source-text context column.

    If term is not provided and term_source exists, use term_source per row.
    """
    if term is None and "term_source" in df.columns:
        return make_kwic_by_term_source(df, context_chars=context_chars)

    if term is None:
        raise ValueError("Please provide term or use make_kwic_by_term_source().")

    term = str(term).strip()

    def extract(text):
        text = "" if pd.isna(text) else str(text)
        pos = text.find(term)

        if pos == -1:
            return text[:context_chars * 2]

        start = max(0, pos - context_chars)
        end = min(len(text), pos + len(term) + context_chars)

        return (
            ("…" if start > 0 else "")
            + text[start:end]
            + ("…" if end < len(text) else "")
        )

    out = df.copy()

    kwic_values = out["source_text"].apply(extract)

    if "kwic_source" in out.columns:
        out = out.drop(columns=["kwic_source"])

    if "category" in out.columns:
        insert_position = out.columns.get_loc("category") + 1
    elif "term_source" in out.columns:
        insert_position = out.columns.get_loc("term_source") + 1
    else:
        insert_position = 0

    out.insert(insert_position, "kwic_source", kwic_values)

    return out

def _normalize_for_alignment_key(value):
    """
    Normalize text for detecting repeated alignment rows.
    """
    if pd.isna(value):
        return ""

    text = str(value)
    text = re.sub(r"\s+", " ", text).strip()

    return text


def _join_unique_nonempty(values, separator="\n---\n"):
    """
    Join unique non-empty values while preserving order.

    Useful when one source segment was repeated because target texts
    were split or merged differently.
    """
    seen = set()
    output = []

    for value in values:
        if pd.isna(value):
            continue

        text = str(value).strip()

        if not text:
            continue

        if text not in seen:
            seen.add(text)
            output.append(text)

    return separator.join(output)



def add_source_occurrence_tracking(df):
    """
    Add occurrence tracking columns without collapsing rows.

    This detects repeated source_text rows for the same term_source and
    section only when they are consecutive in the original input corpus.

    It adds:
    - source_occurrence_id
    - alignment_repeat_index
    - alignment_repeat_count
    - count_as_source_occurrence

    The first row of a repeated source occurrence receives
    count_as_source_occurrence = 1. Continuation rows receive 0.
    """
    if df.empty:
        return df.copy()

    required = ["term_source", "source_text"]

    for col in required:
        if col not in df.columns:
            raise ValueError(f"Column '{col}' not found.")

    out = df.copy().reset_index(drop=True)

    block_ids = []
    previous_key = None
    previous_source_row_number = None
    block_number = 0

    for current_position, row in out.iterrows():
        key = (
            _normalize_for_alignment_key(row.get("term_source", "")),
            _normalize_for_alignment_key(row.get("section", "")),
            _normalize_for_alignment_key(row.get("source_text", "")),
        )

        if "source_row_number" in out.columns and not pd.isna(row.get("source_row_number", pd.NA)):
            try:
                current_source_row_number = int(row.get("source_row_number"))
            except Exception:
                current_source_row_number = current_position
        else:
            current_source_row_number = current_position

        is_consecutive_source_row = (
            previous_source_row_number is not None
            and current_source_row_number == previous_source_row_number + 1
        )

        if key != previous_key or not is_consecutive_source_row:
            block_number += 1

        block_ids.append(block_number)
        previous_key = key
        previous_source_row_number = current_source_row_number

    out["_alignment_block_id"] = block_ids

    out["source_occurrence_id"] = out["_alignment_block_id"].apply(
        lambda x: f"OCC_{int(x):05d}"
    )

    out["alignment_repeat_index"] = (
        out.groupby("_alignment_block_id")
        .cumcount()
        + 1
    )

    out["alignment_repeat_count"] = (
        out.groupby("_alignment_block_id")["_alignment_block_id"]
        .transform("size")
    )

    out["count_as_source_occurrence"] = (
        out["alignment_repeat_index"] == 1
    ).astype(int)

    if "token_hit_count" in out.columns:
        out["counted_token_hits"] = (
            out["token_hit_count"] * out["count_as_source_occurrence"]
        )

    out = out.drop(columns=["_alignment_block_id"])

    return out

def collapse_repeated_source_segments(
    df,
    separator="\n---\n"
):
    """
    Collapse consecutive repeated source_text rows into one source occurrence.

    This is useful for literary translation alignment, where one source
    segment may correspond to multiple target rows because translations
    split or merge sentences differently.

    The function preserves traceability by joining original IDs.
    """
    if df.empty:
        return df.copy()

    tracked = add_source_occurrence_tracking(df)

    group_col = "source_occurrence_id"

    collapsed_rows = []

    for occurrence_id, group in tracked.groupby(group_col, sort=False):
        first = group.iloc[0].copy()

        row = {}

        row["term_source"] = first.get("term_source", "")

        if "category" in group.columns:
            row["category"] = first.get("category", "")
        else:
            row["category"] = ""

        row["source_occurrence_id"] = occurrence_id
        row["alignment_rows_collapsed"] = len(group)

        if "token_hit_count" in group.columns:
          row["token_hit_count"] = int(first.get("token_hit_count", 0) or 0)

        if "id" in group.columns:
            row["id"] = _join_unique_nonempty(
                group["id"],
                separator="; "
            )

        if "section" in group.columns:
            row["section"] = first.get("section", "")

        if "source_row_number" in group.columns:
            row["source_row_number"] = _join_unique_nonempty(
                group["source_row_number"],
                separator="; "
            )

        if "kwic_source" in group.columns:
            row["kwic_source"] = first.get("kwic_source", "")

        if "source_text" in group.columns:
            row["source_text"] = first.get("source_text", "")

        for target_col in TARGET_COLUMNS:
            if target_col in group.columns:
                row[target_col] = _join_unique_nonempty(
                    group[target_col],
                    separator=separator
                )

        # Preserve any extra non-strategy metadata columns.
        protected_cols = set(
            [
                "term_source",
                "category",
                "source_occurrence_id",
                "alignment_rows_collapsed",
                "id",
                "section",
                "kwic_source",
                "source_text",
                "source_row_number",
            ]
            + TARGET_COLUMNS
            + STRATEGY_COLUMNS
            + [
                "alignment_repeat_index",
                "alignment_repeat_count",
                "count_as_source_occurrence",
            ]
        )

        for col in group.columns:
            if col not in protected_cols:
                row[col] = first.get(col, "")

        collapsed_rows.append(row)

    collapsed = pd.DataFrame(collapsed_rows)

    preferred_order = (
        ["term_source", "category", "source_occurrence_id", "alignment_rows_collapsed"]
        + (["token_hit_count"] if "token_hit_count" in collapsed.columns else [])
        + (["kwic_source"] if "kwic_source" in collapsed.columns else [])
        + (["source_row_number"] if "source_row_number" in collapsed.columns else [])
        + ["id", "section", "source_text"]
        + TARGET_COLUMNS
    )

    existing_preferred_order = [
        col for col in preferred_order
        if col in collapsed.columns
    ]

    remaining_cols = [
        col for col in collapsed.columns
        if col not in existing_preferred_order
    ]

    collapsed = collapsed[existing_preferred_order + remaining_cols]

    return collapsed

def search_by_category(
    df,
    categories,
    exact_match=True,
    case_sensitive=False
):
    """
    Search corpus rows by one or more categories.

    Category is optional in the minimal HanParal input schema.
    This function only works for pre-tagged corpora that already contain
    a category column.
    """

    if "category" not in df.columns:
        raise ValueError(
            "Column 'category' not found. Category search only works "
            "with a pre-tagged corpus. Use single_term or multi_term "
            "for a minimal aligned corpus."
        )

    if isinstance(categories, str):
        categories = [categories]

    cleaned_categories = [
        str(category).strip()
        for category in categories
        if str(category).strip()
    ]

    if not cleaned_categories:
        raise ValueError("Please provide at least one category.")

    category_series = df["category"].fillna("").astype(str).str.strip()

    if not case_sensitive:
        category_series_compare = category_series.str.lower()
        categories_compare = [
            category.lower()
            for category in cleaned_categories
        ]
    else:
        category_series_compare = category_series
        categories_compare = cleaned_categories

    if exact_match:
        mask = category_series_compare.isin(categories_compare)
    else:
        mask = pd.Series(False, index=df.index)

        for category in categories_compare:
            mask = mask | category_series_compare.str.contains(
                re.escape(category),
                na=False
            )

    results = df.loc[mask].copy().reset_index(drop=True)

    if "term_source" not in results.columns:
        results.insert(0, "term_source", "")

    if "category" not in results.columns:
        results.insert(1, "category", "")

    return results


def list_categories(df):
    """
    List available categories in the corpus with row counts.

    The category column is optional in HanParal's minimal input schema.
    If it is not present, this returns an empty table instead of raising an error.
    """
    if "category" not in df.columns:
        return pd.DataFrame(columns=["category", "rows"])

    category_summary = (
        df["category"]
        .fillna("")
        .astype(str)
        .str.strip()
        .replace("", pd.NA)
        .dropna()
        .value_counts()
        .reset_index()
    )

    category_summary.columns = ["category", "rows"]

    return category_summary


def create_category_summary(search_results):
    """
    Create a summary table for results grouped by category.
    """
    if search_results.empty:
        return pd.DataFrame(
            columns=["category", "term_source", "occurrences"]
        )

    group_cols = []

    if "category" in search_results.columns:
        group_cols.append("category")

    if "term_source" in search_results.columns:
        group_cols.append("term_source")

    if not group_cols:
        return pd.DataFrame(columns=["category", "term_source", "occurrences"])

    summary = (
        search_results
        .groupby(group_cols, dropna=False)
        .size()
        .reset_index(name="occurrences")
        .sort_values("occurrences", ascending=False)
        .reset_index(drop=True)
    )

    return summary



def create_term_summary(search_results):
    """
    Create a frequency summary for searched terms.

    In the recommended token-level workflow:

    - token_hits:
        number of token-hit rows.
        This is closest to an AntConc hit count.

    - source_occurrences:
        number of unique source passages/segments where the term appears.

    - matching_rows:
        number of rows in the current search_results table.
        In token-level output, this is usually the same as token_hits.
    """
    if search_results.empty:
        return pd.DataFrame(
            columns=[
                "term_source",
                "category",
                "token_hits",
                "source_occurrences",
                "matching_rows",
            ]
        )

    group_cols = []

    if "term_source" in search_results.columns:
        group_cols.append("term_source")

    if "category" in search_results.columns:
        group_cols.append("category")

    if not group_cols:
        return pd.DataFrame(
            columns=[
                "term_source",
                "category",
                "token_hits",
                "source_occurrences",
                "matching_rows",
            ]
        )

    grouped = search_results.groupby(group_cols, dropna=False)

    summary = grouped.size().reset_index(name="matching_rows")

    # Token hits
    if "token_hit_id" in search_results.columns:
        token_hits = (
            grouped["token_hit_id"]
            .count()
            .reset_index(name="token_hits")
        )
    elif "token_hit_count" in search_results.columns:
        token_hits = (
            grouped["token_hit_count"]
            .sum()
            .reset_index(name="token_hits")
        )
    else:
        token_hits = (
            grouped.size()
            .reset_index(name="token_hits")
        )

    # Source occurrences/passages
    if "source_occurrence_id" in search_results.columns:
        source_occurrences = (
            grouped["source_occurrence_id"]
            .nunique()
            .reset_index(name="source_occurrences")
        )
    elif "id" in search_results.columns:
        # Count unique base aligned segment IDs.
        source_occurrences = (
            grouped["id"]
            .nunique()
            .reset_index(name="source_occurrences")
        )
    else:
        source_occurrences = (
            grouped.size()
            .reset_index(name="source_occurrences")
        )

    summary = summary.merge(
        source_occurrences,
        on=group_cols,
        how="left"
    )

    summary = summary.merge(
        token_hits,
        on=group_cols,
        how="left"
    )

    summary = summary[
        group_cols
        + [
            "token_hits",
            "source_occurrences",
            "matching_rows",
        ]
    ]

    summary = (
        summary
        .sort_values("token_hits", ascending=False)
        .reset_index(drop=True)
    )

    return summary


def prepare_annotation_table(df):
    """
    Prepare a token-level table for manual annotation.

    Recommended Phase 1 output:
    one row = one token hit.

    The table includes:
    - term_source: generated from the searched term
    - category: empty by default, with dropdown added during export
    - token-hit metadata
    - KWIC context
    - source and target aligned texts
    - strategy columns
    - notes
    """
    out = df.copy()

    if "term_source" not in out.columns:
        out.insert(0, "term_source", "")

    if "category" not in out.columns:
        term_position = out.columns.get_loc("term_source")
        out.insert(term_position + 1, "category", "")

    for strategy_col in STRATEGY_COLUMNS:
        if strategy_col not in out.columns:
            out[strategy_col] = ""

    if "notes" not in out.columns:
        out["notes"] = ""

    preferred_order = (
        ["term_source", "category"]
        + (["source_occurrence_id"] if "source_occurrence_id" in out.columns else [])
        + (["alignment_rows_collapsed"] if "alignment_rows_collapsed" in out.columns else [])
        + (["source_row_number"] if "source_row_number" in out.columns else [])
        + (["token_hit_id"] if "token_hit_id" in out.columns else [])
        + (["token_hit_index"] if "token_hit_index" in out.columns else [])
        + (["token_hit_count_in_segment"] if "token_hit_count_in_segment" in out.columns else [])
        + (["token_start"] if "token_start" in out.columns else [])
        + (["token_end"] if "token_end" in out.columns else [])
        + (["kwic_source"] if "kwic_source" in out.columns else [])
        + ["id", "section", "source_text"]
        + TARGET_COLUMNS
        + STRATEGY_COLUMNS
        + ["notes"]
    )

    existing_preferred_order = [
        col for col in preferred_order
        if col in out.columns
    ]

    remaining_cols = [
        col for col in out.columns
        if col not in existing_preferred_order
    ]

    out = out[existing_preferred_order + remaining_cols]

    return out

def detect_equivalents(
    df,
    equivalents: Dict[str, Iterable[str]],
    exact_word=False,
    remove_accents=True
):
    """
    Detect whether expected equivalent forms appear in target texts.

    Example:

    expected_equivalents = {
        "target_1": ["menina", "irmã"],
        "target_2": ["girl", "sister"],
        "target_3": ["chica", "hermana"],
    }
    """
    out = df.copy()

    for target_key, forms in equivalents.items():
        target_key = target_key.lower().strip()

        if target_key not in LANGUAGE_COLUMNS:
            raise ValueError(
                "Language keys must be 'target_1', 'target_2', or 'target_3'."
            )

        text_col = LANGUAGE_COLUMNS[target_key]

        forms = [
            str(f).strip()
            for f in forms
            if str(f).strip()
        ]

        matched_all = []

        for _, row in out.iterrows():
            found = [
                f for f in forms
                if contains_form(
                    row[text_col],
                    f,
                    exact_word=exact_word,
                    remove_accents=remove_accents
                )
            ]

            matched_all.append(found)

        out[f"matched_forms_{target_key}"] = [
            ", ".join(x) for x in matched_all
        ]

        out[f"match_{target_key}"] = [
            bool(x) for x in matched_all
        ]

    return out


# ============================================================
# 7. Strategy counting
# ============================================================

def _iter_strategy_values(values):
    """
    Yield non-empty strategy labels from a Series/list.

    HanParal v0.13 uses multiple dropdown slots per target, so each cell
    should contain one strategy. This helper also tolerates old files where
    users manually separated multiple strategies with semicolon.
    """
    for value in values:
        if pd.isna(value):
            continue

        text = str(value).strip()

        if not text:
            continue

        parts = [
            part.strip()
            for part in re.split(r"\s*;\s*", text)
            if part.strip()
        ]

        for part in parts:
            yield part


def has_strategy_columns(df):
    """
    Check whether the file has the expected manual annotation columns.
    """
    return all(col in df.columns for col in STRATEGY_COLUMNS)


def count_strategies(
    df,
    term=None,
    language_labels=None,
    include_percent=True,
):
    """
    Count manually annotated strategy occurrences per target text.

    Supports multiple strategy dropdown slots per target.

    Example:
    strategy_target_1_1 = "tradução linguística"
    strategy_target_1_2 = "eliminação"

    Both are counted for target_1.
    """
    available_strategy_columns = [
        col for col in STRATEGY_COLUMNS
        if col in df.columns
    ]

    if not available_strategy_columns:
        raise ValueError(
            "No strategy columns found. "
            "First export an occurrence table, annotate the strategies, "
            "and then load the annotated file."
        )

    if language_labels is None:
        language_labels = {
            target_col: target_col
            for target_col in TARGET_COLUMNS
        }

    subset = df.copy()

    if term:
        if "term_source" not in subset.columns:
            raise ValueError("Column 'term_source' not found.")
        subset = subset[subset["term_source"] == term]
        if subset.empty:
            raise ValueError(
                f"No rows found for term '{term}'. "
                "Check the term_source column."
            )

    records = []

    for target_col, strategy_cols in STRATEGY_MAP.items():
        target_values = []

        for strategy_col in strategy_cols:
            if strategy_col not in subset.columns:
                continue

            target_values.extend(
                list(_iter_strategy_values(subset[strategy_col]))
            )

        if not target_values:
            continue

        counts = pd.Series(target_values).value_counts()
        total = counts.sum()

        for strategy, count in counts.items():
            record = {
                "language": target_col,
                "language_label": language_labels.get(target_col, target_col),
                "strategy": strategy,
                "count": int(count),
            }

            if include_percent:
                record["percent"] = round((count / total) * 100, 1) if total else 0

            records.append(record)

    if not records:
        return pd.DataFrame(
            columns=[
                "language",
                "language_label",
                "strategy",
                "count",
            ] + (["percent"] if include_percent else [])
        )

    summary = pd.DataFrame(records)

    summary = summary.sort_values(
        ["language", "count", "strategy"],
        ascending=[True, False, True]
    ).reset_index(drop=True)

    return summary

# ============================================================
# 8. Visualization
# ============================================================


def plot_strategy_heatmap(
    df,
    terms=None,
    language_labels=None,
    figsize=None,
    save_path=None,
    annotate=True,
):
    """
    Plot a term × strategy heatmap for each configured target text.

    Supports multiple strategy dropdown slots per target, such as:
    strategy_target_1_1, strategy_target_1_2, strategy_target_1_3.
    """
    if "term_source" not in df.columns:
        raise ValueError("Column 'term_source' not found.")

    if language_labels is None:
        language_labels = {
            target_col: target_col
            for target_col in TARGET_COLUMNS
        }

    available_strategy_map = {
        target_col: [
            strategy_col
            for strategy_col in strategy_cols
            if strategy_col in df.columns
        ]
        for target_col, strategy_cols in STRATEGY_MAP.items()
    }

    available_strategy_map = {
        target_col: strategy_cols
        for target_col, strategy_cols in available_strategy_map.items()
        if strategy_cols
    }

    if not available_strategy_map:
        raise ValueError(
            "No strategy columns found. "
            "Upload or generate an annotated occurrence table first."
        )

    subset = df.copy()

    if terms is None or terms == "":
        terms = (
            subset["term_source"]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace("", pd.NA)
            .dropna()
            .drop_duplicates()
            .tolist()
        )
    else:
        if isinstance(terms, str):
            terms = parse_items_text(terms)
        else:
            terms = [str(term).strip() for term in terms if str(term).strip()]

        subset = subset[subset["term_source"].isin(terms)].copy()

    if subset.empty:
        raise ValueError("No data found for the specified term(s).")

    present_terms = set(subset["term_source"].fillna("").astype(str).str.strip())
    terms = [term for term in terms if term in present_terms]

    if not terms:
        raise ValueError("No valid non-empty terms found for the heatmap.")

    all_strategies = []
    strategy_frequency = {}

    for strategy_cols in available_strategy_map.values():
        for strategy_col in strategy_cols:
            values = list(_iter_strategy_values(subset[strategy_col]))

            for value in values:
                if value not in all_strategies:
                    all_strategies.append(value)
                strategy_frequency[value] = strategy_frequency.get(value, 0) + 1

    if not all_strategies:
        raise ValueError("No annotated strategy values found for the heatmap.")

    all_strategies = sorted(
        all_strategies,
        key=lambda strategy: (-strategy_frequency.get(strategy, 0), strategy)
    )

    matrices = {}

    for target_col, strategy_cols in available_strategy_map.items():
        label = language_labels.get(target_col, target_col)
        matrix = pd.DataFrame(0, index=terms, columns=all_strategies)

        for term in terms:
            term_subset = subset.loc[subset["term_source"] == term]

            values_for_target = []

            for strategy_col in strategy_cols:
                values_for_target.extend(
                    list(_iter_strategy_values(term_subset[strategy_col]))
                )

            counts = pd.Series(values_for_target).value_counts() if values_for_target else pd.Series(dtype=int)

            for strategy, count in counts.items():
                if strategy in matrix.columns:
                    matrix.loc[term, strategy] = int(count)

        matrices[label] = matrix

    n_targets = len(matrices)
    n_terms = len(terms)
    n_strategies = len(all_strategies)

    if figsize is None:
        width = max(10, n_targets * max(4.5, n_strategies * 0.85))
        height = max(5, n_terms * 0.55 + 2.8)
        figsize = (width, height)

    plt.rcParams["font.family"] = ["Noto Sans CJK JP", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False

    fig, axes = plt.subplots(
        1,
        n_targets,
        figsize=figsize,
        sharey=True,
        squeeze=False,
    )

    axes = axes.ravel()
    fig.patch.set_facecolor(HP_BG)

    global_max = max(matrix.values.max() for matrix in matrices.values())
    global_max = global_max if global_max else 1

    for ax, (target_label, matrix) in zip(axes, matrices.items()):
        data = matrix.values.astype(float)

        for row_idx, term in enumerate(matrix.index):
            for col_idx, strategy in enumerate(matrix.columns):
                value = data[row_idx, col_idx]
                norm_value = value / global_max if global_max else 0

                if value == 0:
                    face_color = HP_EMPTY
                    text_color = "#CCBBB5"
                else:
                    rgba = HP_CMAP(norm_value * 0.9 + 0.1)
                    face_color = mcolors.to_hex(rgba)
                    text_color = "#FFFFFF" if norm_value > 0.55 else HP_TEXT

                rectangle = plt.Rectangle(
                    [col_idx, row_idx],
                    1,
                    1,
                    facecolor=face_color,
                    edgecolor=HP_BG,
                    linewidth=1.5,
                    zorder=2,
                )
                ax.add_patch(rectangle)

                if annotate and value > 0:
                    ax.text(
                        col_idx + 0.5,
                        row_idx + 0.5,
                        str(int(value)),
                        ha="center",
                        va="center",
                        color=text_color,
                        fontsize=9,
                        fontweight="bold",
                        zorder=3,
                    )

        ax.set_xlim(0, len(matrix.columns))
        ax.set_ylim(0, len(matrix.index))
        ax.invert_yaxis()

        ax.set_xticks(np.arange(len(matrix.columns)) + 0.5)
        ax.set_xticklabels(
            matrix.columns,
            rotation=45,
            ha="right",
            fontsize=9,
            color=HP_TEXT,
        )

        ax.set_yticks(np.arange(len(matrix.index)) + 0.5)
        ax.set_yticklabels(
            matrix.index,
            fontsize=10,
            color=HP_TEXT,
        )

        ax.set_title(
            target_label,
            fontsize=12,
            fontweight="bold",
            color=HP_TEXT,
            pad=12,
        )

        ax.tick_params(length=0)

        for spine in ax.spines.values():
            spine.set_visible(False)

        ax.set_facecolor(HP_BG)

    title = "Term × strategy heatmap"
    subtitle = (
        "Manual annotation · one row per token hit · "
        f"{len(TARGET_COLUMNS)} target text(s) · "
        f"{N_STRATEGY_SLOTS} strategy slot(s) per target"
    )

    fig.suptitle(
        title,
        fontsize=15,
        fontweight="bold",
        color=HP_TEXT,
        x=0.01,
        ha="left",
        y=1.03,
    )

    fig.text(
        0.01,
        0.99,
        subtitle,
        fontsize=10,
        color=HP_TEXT_MUTED,
        ha="left",
        va="top",
    )

    plt.tight_layout(pad=1.5)

    if save_path:
        fig.savefig(
            save_path,
            dpi=180,
            bbox_inches="tight",
            facecolor=HP_BG,
        )
        print(f"Saved: {save_path}")

    return fig, axes

def plot_strategy_distribution(
    strategy_summary,
    term=None,
    language_labels=None,
    figsize=(11, 6),
    bar_height=0.25,
    save_path=None,
):
    """
    Plot strategy distribution as a horizontal grouped bar chart.

    Parameters
    ----------
    strategy_summary : DataFrame
        Long-format table with columns: language, strategy, count.
    term : str, optional
        The searched source-text term — shown in the title.
    language_labels : dict, optional
        Maps 'target_1'/'target_2'/'target_3' to display names.
        Defaults to neutral labels such as {"target_1": "target_1", "target_2": "target_2"}.
    figsize : tuple
        Figure size in inches.
    bar_height : float
        Height of each individual bar.
    save_path : str or Path, optional
        If given, saves the figure to this path (PNG/PDF).

    Returns
    -------
    matplotlib.axes.Axes
    """
    if strategy_summary.empty:
        raise ValueError("strategy_summary is empty — no strategies to plot.")

    if language_labels is None:
        language_labels = {
            target_col: target_col
            for target_col in TARGET_COLUMNS
        }

    plt.rcParams["font.family"] = ["Noto Sans CJK JP", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False

    languages = [
        lang for lang in TARGET_COLUMNS
        if lang in strategy_summary["language"].values
    ]

    strategies = (
        strategy_summary
        .groupby("strategy")["count"]
        .sum()
        .sort_values(ascending=True)
        .index
        .tolist()
    )

    n_strategies = len(strategies)
    n_langs      = len(languages)
    group_gap    = bar_height * 0.4
    group_height = bar_height * n_langs + group_gap

    fig_height = max(figsize[1], n_strategies * group_height + 1.5)
    fig, ax = plt.subplots(figsize=(figsize[0], fig_height))

    fig.patch.set_facecolor(HP_BG)
    ax.set_facecolor(HP_BG)

    import matplotlib.patches as mpatches

    for lang_idx, lang_key in enumerate(languages):
        lang_data = (
            strategy_summary[strategy_summary["language"] == lang_key]
            .set_index("strategy")["count"]
        )

        counts  = [lang_data.get(s, 0) for s in strategies]
        offsets = [
            i * group_height + lang_idx * bar_height
            for i in range(n_strategies)
        ]

        color = HP_COLORS.get(lang_key, "#8B4A38")
        label = language_labels.get(lang_key, lang_key)

        bars = ax.barh(
            offsets, counts, height=bar_height,
            color=color, label=label, zorder=3,
        )

        for bar, count in zip(bars, counts):
            if count > 0:
                ax.text(
                    bar.get_width() + 0.15,
                    bar.get_y() + bar.get_height() / 2,
                    str(int(count)),
                    va="center", ha="left",
                    fontsize=9, color=HP_TEXT_MUTED,
                )

    tick_positions = [
        i * group_height + (n_langs - 1) * bar_height / 2
        for i in range(n_strategies)
    ]

    ax.set_yticks(tick_positions)
    ax.set_yticklabels(strategies, fontsize=10, color=HP_TEXT)
    ax.set_ylim(-group_gap, n_strategies * group_height)
    ax.set_xlabel("Occurrences", fontsize=10, color=HP_TEXT_MUTED, labelpad=8)
    ax.tick_params(axis="x", colors=HP_TEXT_MUTED, labelsize=9)

    ax.xaxis.grid(True, color=HP_GRID, linewidth=0.8, zorder=0)
    ax.set_axisbelow(True)

    for spine in ["top", "right", "left"]:
        ax.spines[spine].set_visible(False)
    ax.spines["bottom"].set_color(HP_SPINE)
    ax.tick_params(axis="y", length=0)

    title = (
        f"Strategy distribution — {term}"
        if term
        else "Strategy distribution"
    )

    subtitle = f"Manual annotation · source text → {len(TARGET_COLUMNS)} target text(s)"

    ax.set_title(
        title,
        fontsize=13,
        fontweight="bold",
        color=HP_TEXT,
        pad=14,
        loc="left"
    )

    ax.text(
        0,
        1.01,
        subtitle,
        transform=ax.transAxes,
        fontsize=9,
        color=HP_TEXT_MUTED,
        va="bottom"
    )

    legend_patches = [
        mpatches.Patch(
            color=HP_COLORS.get(lang, "#8B4A38"),
            label=language_labels.get(lang, lang)
        )
        for lang in languages
    ]

    ax.legend(
        handles=legend_patches,
        loc="lower right",
        frameon=True,
        framealpha=0.9,
        facecolor=HP_BG,
        edgecolor=HP_SPINE,
        fontsize=10,
        title="Target text",
        title_fontsize=9,
    )

    plt.tight_layout(pad=1.5)

    if save_path:
        fig.savefig(
            save_path,
            dpi=180,
            bbox_inches="tight",
            facecolor=HP_BG
        )
        print(f"Saved: {save_path}")

    return ax



# ============================================================
# 9. Thesis-ready summary tables
# ============================================================

def _clean_nonempty(series):
    """
    Return a stripped string Series with empty values removed.
    """
    return (
        series
        .fillna("")
        .astype(str)
        .str.strip()
        .replace("", pd.NA)
        .dropna()
    )


def create_term_frequency_table(df):
    """
    Create a thesis-ready frequency table by source term and category.

    Output columns:
    - term_source
    - category
    - occurrences
    """
    if "term_source" not in df.columns:
        raise ValueError("Column 'term_source' not found.")

    group_cols = ["term_source"]

    if "category" in df.columns:
        group_cols.append("category")

    table = (
        df
        .groupby(group_cols, dropna=False)
        .size()
        .reset_index(name="occurrences")
        .sort_values(["occurrences", "term_source"], ascending=[False, True])
        .reset_index(drop=True)
    )

    return table


def create_strategy_by_target_table(strategy_summary):
    """
    Create a wide table showing strategy distribution by target text.

    Input must be the long-format table returned by count_strategies():
    language | strategy | count

    Output columns:
    strategy | target_1 | target_2 | ... | total
    """
    required = {"language", "strategy", "count"}

    if strategy_summary is None or strategy_summary.empty:
        return pd.DataFrame(columns=["strategy"] + TARGET_COLUMNS + ["total"])

    if not required.issubset(strategy_summary.columns):
        raise ValueError("strategy_summary must contain: language, strategy, count")

    table = (
        strategy_summary
        .pivot_table(
            index="strategy",
            columns="language",
            values="count",
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
    )

    for target_col in TARGET_COLUMNS:
        if target_col not in table.columns:
            table[target_col] = 0

    ordered_cols = ["strategy"] + TARGET_COLUMNS
    table = table[ordered_cols]
    table["total"] = table[TARGET_COLUMNS].sum(axis=1)

    table = (
        table
        .sort_values("total", ascending=False)
        .reset_index(drop=True)
    )

    return table



def create_strategy_by_category_table(df):
    """
    Create a thesis-ready table of strategy distribution by category.

    Supports multiple strategy dropdown slots per target.

    Output columns:
    category | strategy | target_1 | target_2 | ... | total
    """
    if "category" not in df.columns:
        return pd.DataFrame(columns=["category", "strategy"] + TARGET_COLUMNS + ["total"])

    available_strategy_columns = [
        col for col in STRATEGY_COLUMNS
        if col in df.columns
    ]

    if not available_strategy_columns:
        raise ValueError("No strategy columns found.")

    records = []

    for target_col, strategy_cols in STRATEGY_MAP.items():
        for _, row in df.iterrows():
            category = "" if pd.isna(row["category"]) else str(row["category"]).strip()

            for strategy_col in strategy_cols:
                if strategy_col not in df.columns:
                    continue

                for strategy in _iter_strategy_values([row[strategy_col]]):
                    records.append({
                        "category": category,
                        "strategy": strategy,
                        "language": target_col,
                        "count": 1,
                    })

    if not records:
        return pd.DataFrame(columns=["category", "strategy"] + TARGET_COLUMNS + ["total"])

    long_df = pd.DataFrame(records)

    table = (
        long_df
        .pivot_table(
            index=["category", "strategy"],
            columns="language",
            values="count",
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
    )

    for target_col in TARGET_COLUMNS:
        if target_col not in table.columns:
            table[target_col] = 0

    ordered_cols = ["category", "strategy"] + TARGET_COLUMNS
    table = table[ordered_cols]
    table["total"] = table[TARGET_COLUMNS].sum(axis=1)

    table = (
        table
        .sort_values(["category", "total", "strategy"], ascending=[True, False, True])
        .reset_index(drop=True)
    )

    return table


def create_strategy_by_term_table(df):
    """
    Create a thesis-ready table summarizing the main strategy per term.

    Supports multiple strategy dropdown slots per target.

    Output columns:
    term_source | category | occurrences |
    target_1_main_strategy | target_1_strategy_counts | ...
    """
    if "term_source" not in df.columns:
        raise ValueError("Column 'term_source' not found.")

    available_strategy_columns = [
        col for col in STRATEGY_COLUMNS
        if col in df.columns
    ]

    if not available_strategy_columns:
        raise ValueError("No strategy columns found.")

    group_cols = ["term_source"]

    if "category" in df.columns:
        group_cols.append("category")

    rows = []

    for group_values, group in df.groupby(group_cols, dropna=False):
        if not isinstance(group_values, tuple):
            group_values = (group_values,)

        row_data = {
            col: value
            for col, value in zip(group_cols, group_values)
        }

        row_data["occurrences"] = len(group)

        for target_col, strategy_cols in STRATEGY_MAP.items():
            values = []

            for strategy_col in strategy_cols:
                if strategy_col in group.columns:
                    values.extend(list(_iter_strategy_values(group[strategy_col])))

            if not values:
                row_data[f"{target_col}_main_strategy"] = ""
                row_data[f"{target_col}_strategy_counts"] = ""
            else:
                counts = pd.Series(values).value_counts()
                row_data[f"{target_col}_main_strategy"] = counts.index[0]
                row_data[f"{target_col}_strategy_counts"] = "; ".join(
                    f"{strategy} ({count})"
                    for strategy, count in counts.items()
                )

        rows.append(row_data)

    table = pd.DataFrame(rows)

    if not table.empty:
        table = table.sort_values(
            ["occurrences", "term_source"],
            ascending=[False, True]
        ).reset_index(drop=True)

    return table

def create_thesis_overview_table(df, strategy_summary=None):
    """
    Create a compact overview table for the thesis summary file.
    """
    total_strategy_cells = 0
    empty_strategy_cells = 0

    available_strategy_columns = [
        col for col in STRATEGY_COLUMNS
        if col in df.columns
    ]

    for col in available_strategy_columns:
        total_strategy_cells += len(df[col])
        empty_strategy_cells += (
            df[col]
            .fillna("")
            .astype(str)
            .str.strip()
            .eq("")
            .sum()
        )

    unique_strategies = 0

    if strategy_summary is not None and not strategy_summary.empty:
        unique_strategies = strategy_summary["strategy"].nunique()

    metrics = [
        {"metric": "rows_analyzed", "value": len(df)},
        {"metric": "target_texts", "value": len(TARGET_COLUMNS)},
        {"metric": "unique_terms", "value": df["term_source"].nunique() if "term_source" in df.columns else 0},
        {"metric": "unique_categories", "value": df["category"].nunique() if "category" in df.columns else 0},
        {"metric": "unique_strategies", "value": unique_strategies},
        {"metric": "strategy_cells", "value": total_strategy_cells},
        {"metric": "empty_strategy_cells", "value": int(empty_strategy_cells)},
    ]

    return pd.DataFrame(metrics)


def generate_thesis_summary_tables(df, strategy_summary=None):
    """
    Generate a dictionary of thesis-ready summary tables.

    The tables are language-neutral and adapt to the configured number
    of target texts.
    """
    if strategy_summary is None:
        strategy_summary = count_strategies(df)

    tables = {
        "thesis_overview": create_thesis_overview_table(df, strategy_summary),
        "term_frequency": create_term_frequency_table(df),
        "strategy_by_target": create_strategy_by_target_table(strategy_summary),
        "strategy_by_category": create_strategy_by_category_table(df),
        "strategy_by_term": create_strategy_by_term_table(df),
    }

    return tables


def display_thesis_summary_tables(thesis_tables):
    """
    Display thesis-ready summary tables in Colab.
    """
    for name, table in thesis_tables.items():
        print(name)
        display(table)


def export_thesis_summary_tables(excel_path, thesis_tables):
    """
    Export thesis-ready summary tables to an Excel workbook.
    """
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        for sheet_name, table in thesis_tables.items():
            safe_sheet_name = str(sheet_name)[:31]
            table.to_excel(
                writer,
                index=False,
                sheet_name=safe_sheet_name
            )

    format_hanparal_workbook(excel_path)

    return excel_path

# ============================================================
# 9. Export
# ============================================================

def export_excel(
    path,
    search_results=None,
    annotation_table=None,
    equivalent_results=None,
    strategy_summary=None,
    term_summary=None,
    category_summary=None,
    thesis_summary_tables=None
):
    """
    Export one or more HanParal output tables to Excel.
    """
    with pd.ExcelWriter(path, engine="openpyxl") as writer:

        if search_results is not None:
            search_results.to_excel(
                writer,
                index=False,
                sheet_name="search_results"
            )

        if annotation_table is not None:
            annotation_table.to_excel(
                writer,
                index=False,
                sheet_name="annotation_table"
            )

        if equivalent_results is not None:
            equivalent_results.to_excel(
                writer,
                index=False,
                sheet_name="equivalent_detection"
            )

        if strategy_summary is not None:
            strategy_summary.to_excel(
                writer,
                index=False,
                sheet_name="strategy_summary"
            )

        if term_summary is not None:
            term_summary.to_excel(
                writer,
                index=False,
                sheet_name="term_summary"
            )

        if category_summary is not None:
            category_summary.to_excel(
                writer,
                index=False,
                sheet_name="category_summary"
            )

        if thesis_summary_tables is not None:
            for sheet_name, table in thesis_summary_tables.items():
                safe_sheet_name = str(sheet_name)[:31]
                table.to_excel(
                    writer,
                    index=False,
                    sheet_name=safe_sheet_name
                )

    return path

# ============================================================
# 10. Excel dropdowns for strategy/category annotation
# ============================================================

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


# Strategy labels can be customized according to the research methodology.
STRATEGY_OPTIONS = [
    "repetição",
    "adaptação ortográfica",
    "tradução linguística",
    "glosa intratextual",
    "glosa extratextual",
    "sinonímia",
    "universalização limitada",
    "universalização absoluta",
    "naturalização",
    "eliminação",
    "criação autônoma",
]



def add_strategy_dropdowns(
    excel_path,
    sheet_name="annotation_table",
    strategy_options=None,
    strategy_columns=None,
    use_inline_list=True
):
    """
    Add dropdown menus to all strategy annotation columns.

    HanParal uses multiple strategy slots per target, for example:
    strategy_target_1_1, strategy_target_1_2, strategy_target_1_3.

    Each slot receives the same dropdown.

    By default, this uses an inline Excel list because it is more likely
    to survive Google Sheets import than a validation pointing to another
    worksheet range. A visible sheet called strategy_options is still
    created as a fallback/reference list.
    """
    from openpyxl import load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment

    if strategy_options is None:
        strategy_options = STRATEGY_OPTIONS

    if strategy_columns is None:
        strategy_columns = STRATEGY_COLUMNS

    strategy_options = [
        str(option).strip()
        for option in strategy_options
        if str(option).strip()
    ]

    if not strategy_options:
        strategy_options = [
            "Repetição",
            "Adaptação ortográfica",
            "Tradução linguística",
            "Glosa extratextual",
            "Glosa intratextual",
            "Sinônimos",
            "Universalização limitada",
            "Universalização absoluta",
            "Naturalização",
            "Eliminação",
            "Criação autônoma",
            "Compensação",
            "Deslocação",
            "Atenuação",
        ]

    wb = load_workbook(excel_path)

    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. "
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb[sheet_name]

    headers = {
        cell.value: cell.column
        for cell in ws[1]
        if cell.value is not None
    }

    available_strategy_columns = [
        col for col in strategy_columns
        if col in headers
    ]

    if not available_strategy_columns:
        raise ValueError(
            "No strategy columns found in annotation table. "
            "Expected columns such as strategy_target_1_1."
        )

    # Visible fallback/options sheet.
    options_sheet_name = "strategy_options"

    if options_sheet_name in wb.sheetnames:
        options_ws = wb[options_sheet_name]
        options_ws.delete_rows(1, options_ws.max_row)
    else:
        options_ws = wb.create_sheet(options_sheet_name)

    options_ws.sheet_state = "visible"

    options_ws["A1"] = "strategy_options"
    options_ws["A1"].font = Font(bold=True, color="FFFFFF")
    options_ws["A1"].fill = PatternFill(
        start_color="9A4F3B",
        end_color="9A4F3B",
        fill_type="solid"
    )
    options_ws["A1"].alignment = Alignment(horizontal="center")

    for row_number, option in enumerate(strategy_options, start=2):
        options_ws.cell(row=row_number, column=1, value=option)

    options_ws.column_dimensions["A"].width = 34

    if use_inline_list:
        formula = _excel_inline_list_formula(strategy_options)
    else:
        formula = f"={options_sheet_name}!$A$2:$A${len(strategy_options) + 1}"

    for strategy_col in available_strategy_columns:
        col_index = headers[strategy_col]
        col_letter = get_column_letter(col_index)

        data_validation = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True
        )

        data_validation.prompt = (
            "Choose one translation strategy for this slot. "
            "Use the next strategy slot if more than one strategy applies."
        )
        data_validation.promptTitle = "Strategy annotation"
        data_validation.showErrorMessage = False

        ws.add_data_validation(data_validation)
        data_validation.add(f"{col_letter}2:{col_letter}{ws.max_row}")

    wb.save(excel_path)

    return excel_path

# ============================================================
# 11. Excel formatting
# ============================================================

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# HanParal brand palette — Sienna Tijolo
# ============================================================

# Header
HP_HEADER_BG      = "8B4A38"   # sienna tijolo — cabeçalho
HP_HEADER_FONT    = "FFFFFF"   # branco

# Row fills (alternating)
HP_ROW_A          = "F5EBE7"   # blush sienna — linhas ímpares
HP_ROW_B          = "FFF9F7"   # creme quase branco — linhas pares

# Functional columns
HP_STRATEGY_FILL  = "F0DDD7"   # sienna suave — estratégia preenchida
HP_STRATEGY_EMPTY = "FEF3C7"   # âmbar — estratégia vazia (alerta)
HP_NOTES_FILL     = "F0E6FF"   # lavanda — coluna de notas
HP_MATCH_TRUE     = "D4F0E4"   # menta — match encontrado
HP_MATCH_FALSE    = "F0DDD7"   # sienna suave — sem match

# Term highlight
HP_TERM_FONT      = "8B4A38"   # sienna — term_source em destaque

# Borders
HP_BORDER_COLOR   = "E8C8BC"   # borda sienna suave

# Font colors for functional cells
HP_STRATEGY_TEXT  = "6A2A18"
HP_EMPTY_TEXT     = "7A5000"
HP_NOTES_TEXT     = "4A2070"
HP_MATCH_TEXT     = "1A5C3A"




def _excel_inline_list_formula(options):
    """
    Build an Excel-compatible inline list formula for data validation.

    This is often more portable to Google Sheets than a validation
    pointing to a range in another sheet.
    """
    cleaned = []

    for option in options:
        text = str(option).strip()
        if not text:
            continue

        # Excel inline list formulas cannot contain unescaped double quotes.
        text = text.replace('"', "'")
        cleaned.append(text)

    if not cleaned:
        cleaned = [
            "Category 1",
            "Category 2",
            "Category 3",
            "Category 4",
            "Category 5",
        ]

    return '"' + ','.join(cleaned) + '"'


def add_category_dropdown(
    excel_path,
    sheet_name="annotation_table",
    category_options=None,
    use_inline_list=True
):
    """
    Add a dropdown to the category column.

    By default, this uses an inline Excel list such as:
    "Category 1,Category 2,Category 3"

    This is usually more reliable when an .xlsx file is imported into
    Google Sheets than a validation that points to a hidden/range sheet.

    A visible sheet called category_options is also created as a fallback,
    so researchers can recreate the dropdown manually in Google Sheets if
    Google does not preserve imported Excel validations.
    """
    from openpyxl import load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment

    if category_options is None:
        category_options = [
            "Category 1",
            "Category 2",
            "Category 3",
            "Category 4",
            "Category 5",
        ]

    category_options = [
        str(option).strip()
        for option in category_options
        if str(option).strip()
    ]

    if not category_options:
        category_options = [
            "Category 1",
            "Category 2",
            "Category 3",
            "Category 4",
            "Category 5",
        ]

    wb = load_workbook(excel_path)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")

    ws = wb[sheet_name]

    headers = {
        cell.value: cell.column
        for cell in ws[1]
        if cell.value is not None
    }

    if "category" not in headers:
        wb.save(excel_path)
        return excel_path

    # Visible fallback/options sheet.
    options_sheet_name = "category_options"

    if options_sheet_name in wb.sheetnames:
        options_ws = wb[options_sheet_name]
        options_ws.delete_rows(1, options_ws.max_row)
    else:
        options_ws = wb.create_sheet(options_sheet_name)

    options_ws.sheet_state = "visible"

    options_ws["A1"] = "category_options"
    options_ws["A1"].font = Font(bold=True, color="FFFFFF")
    options_ws["A1"].fill = PatternFill(
        start_color="9A4F3B",
        end_color="9A4F3B",
        fill_type="solid"
    )
    options_ws["A1"].alignment = Alignment(horizontal="center")

    for row_number, option in enumerate(category_options, start=2):
        options_ws.cell(row=row_number, column=1, value=option)

    options_ws.column_dimensions["A"].width = 32

    category_col = headers["category"]
    category_col_letter = get_column_letter(category_col)

    if use_inline_list:
        formula = _excel_inline_list_formula(category_options)
    else:
        formula = f"={options_sheet_name}!$A$2:$A${len(category_options) + 1}"

    data_validation = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True
    )

    data_validation.promptTitle = "Category annotation"
    data_validation.prompt = "Choose a category for this token hit."
    data_validation.showErrorMessage = False

    ws.add_data_validation(data_validation)
    data_validation.add(f"{category_col_letter}2:{category_col_letter}{ws.max_row}")

    wb.save(excel_path)

    return excel_path

def _make_border():
    side = Side(style="thin", color=HP_BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


def _make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def format_hanparal_workbook(excel_path):
    """
    Apply HanParal visual formatting — Sienna Tijolo palette.

    Works on all sheets produced by export_excel():
    search_results, annotation_table, equivalent_detection, strategy_summary.
    """
    wb = load_workbook(excel_path)
    border = _make_border()

    for ws in wb.worksheets:
        if ws.sheet_state == "hidden":
            continue

        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        max_row = ws.max_row
        max_col = ws.max_column

        if max_row < 1 or max_col < 1:
            continue

        # ── Header row ──────────────────────────────────────────
        for cell in ws[1]:
            cell.fill    = _make_fill(HP_HEADER_BG)
            cell.font    = Font(color=HP_HEADER_FONT, bold=True, size=11)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = border

        ws.row_dimensions[1].height = 30

        # Header index
        headers = {
            cell.value: cell.column
            for cell in ws[1]
            if cell.value is not None
        }

        # ── Data rows ────────────────────────────────────────────
        for row_idx in range(2, max_row + 1):
            row_fill = HP_ROW_A if row_idx % 2 == 0 else HP_ROW_B

            for cell in ws[row_idx]:
                cell.fill      = _make_fill(row_fill)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border    = border

            ws.row_dimensions[row_idx].height = 50

        # ── Column-specific styling ───────────────────────────────
        for header, col_idx in headers.items():
            col_letter = get_column_letter(col_idx)

            # Width defaults
            width = 18
            if header in ("id", "section", "category", "token_hit_index", "token_hit_count_in_segment", "token_start", "token_end", "alignment_rows_collapsed"):
                width = 13
            elif header == "term_source":
                width = 20
            elif header == "token_hit_id":
                width = 28
            elif header == "source_occurrence_id":
                width = 20
            elif header == "kwic_source":
                width = 48
            elif header == "source_text" or header in TARGET_COLUMNS:
                width = 56
            elif header.startswith("matched_forms"):
                width = 30
            elif header.startswith("match_"):
                width = 14
            elif header.startswith("strategy_"):
                width = 26
            elif header == "notes":
                width = 36
            elif header in ("language", "strategy", "count"):
                width = 24

            ws.column_dimensions[col_letter].width = width

            # term_source → sienna font
            if header == "term_source":
                for row_idx in range(2, max_row + 1):
                    cell = ws[f"{col_letter}{row_idx}"]
                    cell.font = Font(color=HP_TERM_FONT, bold=True, size=11)
                    cell.alignment = Alignment(
                        horizontal="center", vertical="top", wrap_text=True
                    )

            # id / section / category → centered
            elif header in ("id", "section", "category", "token_hit_index", "token_hit_count_in_segment", "token_start", "token_end", "alignment_rows_collapsed"):
                for row_idx in range(2, max_row + 1):
                    ws[f"{col_letter}{row_idx}"].alignment = Alignment(
                        horizontal="center", vertical="top", wrap_text=True
                    )

            # strategy columns → color by fill state
            elif header.startswith("strategy_"):
                for row_idx in range(2, max_row + 1):
                    cell = ws[f"{col_letter}{row_idx}"]
                    is_empty = (
                        cell.value is None
                        or str(cell.value).strip() == ""
                    )
                    if is_empty:
                        cell.fill = _make_fill(HP_STRATEGY_EMPTY)
                        cell.font = Font(
                            color=HP_EMPTY_TEXT, italic=True, size=10
                        )
                    else:
                        cell.fill = _make_fill(HP_STRATEGY_FILL)
                        cell.font = Font(color=HP_STRATEGY_TEXT, size=11)

            # notes column → lavanda
            elif header == "notes":
                for row_idx in range(2, max_row + 1):
                    cell = ws[f"{col_letter}{row_idx}"]
                    cell.fill = _make_fill(HP_NOTES_FILL)
                    cell.font = Font(color=HP_NOTES_TEXT, size=11)

            # match_ columns → menta/sienna + centered
            elif header.startswith("match_") and not header.startswith("matched_forms"):
                for row_idx in range(2, max_row + 1):
                    cell = ws[f"{col_letter}{row_idx}"]
                    is_true = str(cell.value).strip().lower() == "true"
                    cell.fill = _make_fill(
                        HP_MATCH_TRUE if is_true else HP_MATCH_FALSE
                    )
                    cell.font = Font(
                        color=HP_MATCH_TEXT if is_true else HP_STRATEGY_TEXT,
                        size=11
                    )
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )

        # ── Strategy summary special formatting ──────────────────
        if ws.title == "strategy_summary" and {"language", "strategy", "count"}.issubset(headers):
            count_col = get_column_letter(headers["count"])
            for row_idx in range(2, max_row + 1):
                ws[f"{count_col}{row_idx}"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            ws.column_dimensions[count_col].width = 12

    wb.save(excel_path)
    return excel_path

# ============================================================
# 12. Excel chart export
# ============================================================

from openpyxl.chart import BarChart, Reference


__all__ = [
    "configure_target_columns",
    "parse_items_text",
    "make_safe_filename_fragment",
    "validate_base_columns",
    "load_corpus",
    "normalize_text",
    "contains_form",
    "search_source_term",
    "search_multiple_source_terms",
    "count_term_hits_in_text",
    "add_token_hit_counts",
    "expand_search_results_to_token_hits",
    "make_kwic_by_term_source",
    "make_kwic_by_searched_term",
    "make_kwic",
    "add_source_occurrence_tracking",
    "collapse_repeated_source_segments",
    "search_by_category",
    "list_categories",
    "create_category_summary",
    "create_term_summary",
    "prepare_annotation_table",
    "detect_equivalents",
    "has_strategy_columns",
    "count_strategies",
    "plot_strategy_heatmap",
    "plot_strategy_distribution",
    "create_term_frequency_table",
    "create_strategy_by_target_table",
    "create_strategy_by_category_table",
    "create_strategy_by_term_table",
    "create_thesis_overview_table",
    "generate_thesis_summary_tables",
    "display_thesis_summary_tables",
    "export_thesis_summary_tables",
    "export_excel",
    "add_strategy_dropdowns",
    "add_category_dropdown",
    "format_hanparal_workbook",
    "add_strategy_distribution_chart",
    "generate_corpus_health_report",
    "display_health_report",
    "export_health_report",
]




def add_strategy_distribution_chart(
    excel_path,
    summary_sheet="strategy_summary",
    chart_sheet="strategy_distribution"
):
    """
    Add a native Excel bar chart showing strategy distribution.

    It expects the strategy_summary sheet to have this long-format structure:

    language | strategy | count

    The function creates a new sheet with a pivot-style table and chart.
    """
    wb = load_workbook(excel_path)

    if summary_sheet not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{summary_sheet}' not found. "
            f"Available sheets: {wb.sheetnames}"
        )

    ws_summary = wb[summary_sheet]

    # Read summary data
    rows = list(ws_summary.iter_rows(values_only=True))

    if len(rows) < 2:
        raise ValueError("strategy_summary is empty. No chart can be created.")

    headers = list(rows[0])

    required_headers = ["language", "strategy", "count"]

    for header in required_headers:
        if header not in headers:
            raise ValueError(
                f"Column '{header}' not found in strategy_summary."
            )

    language_idx = headers.index("language")
    strategy_idx = headers.index("strategy")
    count_idx = headers.index("count")

    records = []

    for row in rows[1:]:
        language = row[language_idx]
        strategy = row[strategy_idx]
        count = row[count_idx]

        if language is None or strategy is None or count is None:
            continue

        records.append({
            "language": str(language),
            "strategy": str(strategy),
            "count": int(count),
        })

    if not records:
        raise ValueError("No valid strategy records found.")

    # Build pivot-style data
    strategies = sorted(set(record["strategy"] for record in records))
    languages = sorted(set(record["language"] for record in records))

    pivot = {
        strategy: {language: 0 for language in languages}
        for strategy in strategies
    }

    for record in records:
        pivot[record["strategy"]][record["language"]] += record["count"]

    # Replace old chart sheet if it exists
    if chart_sheet in wb.sheetnames:
        del wb[chart_sheet]

    ws_chart = wb.create_sheet(chart_sheet)

    # Write title
    ws_chart["A1"] = "Strategy distribution"
    ws_chart["A1"].font = Font(bold=True, size=14)

    # Write pivot table header
    ws_chart["A3"] = "strategy"

    for col_idx, language in enumerate(languages, start=2):
        ws_chart.cell(row=3, column=col_idx, value=language)

    # Write pivot table data
    for row_idx, strategy in enumerate(strategies, start=4):
        ws_chart.cell(row=row_idx, column=1, value=strategy)

        for col_idx, language in enumerate(languages, start=2):
            ws_chart.cell(
                row=row_idx,
                column=col_idx,
                value=pivot[strategy][language]
            )

    # Create native Excel bar chart
    chart = BarChart()
    chart.type = "bar"       # horizontal bar chart
    chart.style = 10
    chart.title = "Strategy distribution by target text"
    chart.y_axis.title = "Strategy"
    chart.x_axis.title = "Count"

    data = Reference(
        ws_chart,
        min_col=2,
        max_col=1 + len(languages),
        min_row=3,
        max_row=3 + len(strategies)
    )

    categories = Reference(
        ws_chart,
        min_col=1,
        min_row=4,
        max_row=3 + len(strategies)
    )

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart.height = 12
    chart.width = 22

    ws_chart.add_chart(chart, "A8")

    # Basic formatting
    ws_chart.freeze_panes = "A4"
    ws_chart.sheet_view.showGridLines = False

    for cell in ws_chart[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(horizontal="center")

    for col_idx in range(1, 2 + len(languages)):
        col_letter = get_column_letter(col_idx)
        ws_chart.column_dimensions[col_letter].width = 24

    wb.save(excel_path)

    return excel_path

# ============================================================
# 13. Corpus health report
# ============================================================

def _blank_mask(series):
    """
    Return True for empty, NaN, or whitespace-only cells.
    """
    return series.fillna("").astype(str).str.strip().eq("")


def generate_corpus_health_report(
    df,
    check_strategies=False,
    strategy_options=None
):
    """
    Generate a corpus health report.

    Phase 1 use:
        generate_corpus_health_report(corpus)

    Phase 2 use after manual annotation:
        generate_corpus_health_report(annotated, check_strategies=True)

    The report is language-neutral and checks:
    - missing required values
    - duplicate IDs
    - empty source/target cells
    - term_source not found inside source_text
    - duplicate source_text rows
    - optional strategy annotation problems
    """

    if strategy_options is None and "STRATEGY_OPTIONS" in globals():
        strategy_options = STRATEGY_OPTIONS

    report = {}
    summary_records = []

    def add_summary(check, status, issue_count, recommendation):
        summary_records.append({
            "check": check,
            "status": status,
            "issue_count": int(issue_count),
            "recommendation": recommendation,
        })

    total_rows = len(df)

    # ------------------------------------------------------------
    # 1. Basic overview
    # ------------------------------------------------------------

    overview = pd.DataFrame([
        {"metric": "total_rows", "value": total_rows},
        {"metric": "total_columns", "value": len(df.columns)},
        {"metric": "base_columns_present", "value": all(col in df.columns for col in BASE_COLUMNS)},
        {"metric": "strategy_columns_present", "value": all(col in df.columns for col in STRATEGY_COLUMNS)},
    ])

    report["overview"] = overview

    # ------------------------------------------------------------
    # 2. Missing base columns
    # ------------------------------------------------------------

    missing_base_columns = [
        col for col in BASE_COLUMNS
        if col not in df.columns
    ]

    missing_base_columns_df = pd.DataFrame(
        {"missing_base_column": missing_base_columns}
    )

    report["missing_base_columns"] = missing_base_columns_df

    add_summary(
        check="missing_base_columns",
        status="ERROR" if missing_base_columns else "OK",
        issue_count=len(missing_base_columns),
        recommendation=(
            "Add the missing base columns before running HanParal."
            if missing_base_columns
            else "All required base columns are present."
        )
    )

    # If base columns are missing, stop deeper checks safely
    available_base_columns = [
        col for col in BASE_COLUMNS
        if col in df.columns
    ]

    # ------------------------------------------------------------
    # 3. Empty values in base columns
    # ------------------------------------------------------------

    empty_base_records = []

    for col in available_base_columns:
        empty_count = _blank_mask(df[col]).sum()
        empty_base_records.append({
            "column": col,
            "empty_cells": int(empty_count),
            "percent_empty": round((empty_count / total_rows) * 100, 2) if total_rows else 0,
        })

    empty_base_df = pd.DataFrame(empty_base_records)
    report["empty_base_cells_by_column"] = empty_base_df

    empty_base_total = empty_base_df["empty_cells"].sum() if not empty_base_df.empty else 0

    add_summary(
        check="empty_base_cells",
        status="WARNING" if empty_base_total else "OK",
        issue_count=empty_base_total,
        recommendation=(
            "Review empty cells in required corpus columns."
            if empty_base_total
            else "No empty cells found in required base columns."
        )
    )

    # Rows with any empty base cell
    if available_base_columns:
        any_empty_base = pd.Series(False, index=df.index)

        for col in available_base_columns:
            any_empty_base = any_empty_base | _blank_mask(df[col])

        rows_with_empty_base = df.loc[any_empty_base].copy()
    else:
        rows_with_empty_base = pd.DataFrame()

    report["rows_with_empty_base_cells"] = rows_with_empty_base

    # ------------------------------------------------------------
    # 4. Duplicate IDs
    # ------------------------------------------------------------

    if "id" in df.columns:
        duplicate_id_rows = df[df["id"].duplicated(keep=False)].copy()
    else:
        duplicate_id_rows = pd.DataFrame()

    report["duplicate_ids"] = duplicate_id_rows

    add_summary(
        check="duplicate_ids",
        status="WARNING" if len(duplicate_id_rows) else "OK",
        issue_count=len(duplicate_id_rows),
        recommendation=(
            "Check whether duplicated IDs are intentional."
            if len(duplicate_id_rows)
            else "No duplicated IDs found."
        )
    )

    # ------------------------------------------------------------
    # 5. Empty source/target text cells
    # ------------------------------------------------------------

    text_columns = [
        col for col in ["source_text"] + TARGET_COLUMNS
        if col in df.columns
]

    if text_columns:
        empty_text_mask = pd.Series(False, index=df.index)

        for col in text_columns:
            empty_text_mask = empty_text_mask | _blank_mask(df[col])

        rows_with_empty_text = df.loc[empty_text_mask].copy()
    else:
        rows_with_empty_text = pd.DataFrame()

    report["rows_with_empty_text_cells"] = rows_with_empty_text

    add_summary(
        check="empty_source_or_target_text",
        status="WARNING" if len(rows_with_empty_text) else "OK",
        issue_count=len(rows_with_empty_text),
        recommendation=(
            "Review rows where source_text or target texts are empty."
            if len(rows_with_empty_text)
            else "No empty source/target text cells found."
        )
    )

    # ------------------------------------------------------------
    # 6. term_source not found inside source_text
    # ------------------------------------------------------------

    if "term_source" in df.columns and "source_text" in df.columns:
        term_not_found_mask = []

        for _, row in df.iterrows():
            term = "" if pd.isna(row["term_source"]) else str(row["term_source"]).strip()
            source = "" if pd.isna(row["source_text"]) else str(row["source_text"])

            if not term or not source:
                term_not_found_mask.append(False)
            else:
                term_not_found_mask.append(term not in source)

        term_not_found_df = df.loc[term_not_found_mask].copy()
    else:
        term_not_found_df = pd.DataFrame()

    report["term_source_not_found_in_source_text"] = term_not_found_df

    add_summary(
        check="term_source_not_found_in_source_text",
        status="WARNING" if len(term_not_found_df) else "OK",
        issue_count=len(term_not_found_df),
        recommendation=(
            "Check whether term_source is written differently from the form in source_text."
            if len(term_not_found_df)
            else "Every non-empty term_source was found inside source_text."
        )
    )

    # ------------------------------------------------------------
    # 7. Duplicate source_text rows
    # ------------------------------------------------------------

    if "source_text" in df.columns:
        non_empty_source = ~_blank_mask(df["source_text"])
        duplicate_source_rows = df[
            non_empty_source & df["source_text"].duplicated(keep=False)
        ].copy()
    else:
        duplicate_source_rows = pd.DataFrame()

    report["duplicate_source_text"] = duplicate_source_rows

    add_summary(
        check="duplicate_source_text",
        status="WARNING" if len(duplicate_source_rows) else "OK",
        issue_count=len(duplicate_source_rows),
        recommendation=(
            "Check whether repeated source_text rows are intentional."
            if len(duplicate_source_rows)
            else "No duplicated source_text rows found."
        )
    )

    # ------------------------------------------------------------
    # 8. Term distribution
    # ------------------------------------------------------------

    if "term_source" in df.columns:
        term_distribution = (
            df["term_source"]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace("", pd.NA)
            .dropna()
            .value_counts()
            .reset_index()
        )

        term_distribution.columns = ["term_source", "count"]
    else:
        term_distribution = pd.DataFrame(columns=["term_source", "count"])

    report["term_distribution"] = term_distribution

    # ------------------------------------------------------------
    # 9. Category distribution
    # ------------------------------------------------------------

    if "category" in df.columns:
        category_distribution = (
            df["category"]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace("", pd.NA)
            .dropna()
            .value_counts()
            .reset_index()
        )

        category_distribution.columns = ["category", "count"]
    else:
        category_distribution = pd.DataFrame(columns=["category", "count"])

    report["category_distribution"] = category_distribution

    # ------------------------------------------------------------
    # 10. Optional strategy checks
    # ------------------------------------------------------------

    if check_strategies:
        missing_strategy_columns = [
            col for col in STRATEGY_COLUMNS
            if col not in df.columns
        ]

        report["missing_strategy_columns"] = pd.DataFrame({
            "missing_strategy_column": missing_strategy_columns
        })

        add_summary(
            check="missing_strategy_columns",
            status="ERROR" if missing_strategy_columns else "OK",
            issue_count=len(missing_strategy_columns),
            recommendation=(
                "Upload the annotation table with strategy columns."
                if missing_strategy_columns
                else "All strategy columns are present."
            )
        )

        available_strategy_columns = [
            col for col in STRATEGY_COLUMNS
            if col in df.columns
        ]

        # Empty strategy cells
        empty_strategy_records = []

        for col in available_strategy_columns:
            empty_count = _blank_mask(df[col]).sum()
            empty_strategy_records.append({
                "strategy_column": col,
                "empty_cells": int(empty_count),
                "percent_empty": round((empty_count / total_rows) * 100, 2) if total_rows else 0,
            })

        empty_strategy_df = pd.DataFrame(empty_strategy_records)
        report["empty_strategy_cells_by_column"] = empty_strategy_df

        empty_strategy_total = (
            empty_strategy_df["empty_cells"].sum()
            if not empty_strategy_df.empty
            else 0
        )

        add_summary(
            check="empty_strategy_cells",
            status="WARNING" if empty_strategy_total else "OK",
            issue_count=empty_strategy_total,
            recommendation=(
                "Fill empty strategy cells before final counting."
                if empty_strategy_total
                else "No empty strategy cells found."
            )
        )

        # Invalid strategy labels
        invalid_records = []

        if strategy_options is not None:
            valid_options = set(strategy_options)

            for col in available_strategy_columns:
                for idx, value in df[col].items():
                    value_clean = "" if pd.isna(value) else str(value).strip()

                    if value_clean and value_clean not in valid_options:
                        invalid_records.append({
                            "row_index": idx,
                            "id": df.loc[idx, "id"] if "id" in df.columns else "",
                            "strategy_column": col,
                            "invalid_value": value_clean,
                        })

        invalid_strategy_df = pd.DataFrame(
            invalid_records,
            columns=["row_index", "id", "strategy_column", "invalid_value"]
        )

        report["invalid_strategy_labels"] = invalid_strategy_df

        add_summary(
            check="invalid_strategy_labels",
            status="WARNING" if len(invalid_strategy_df) else "OK",
            issue_count=len(invalid_strategy_df),
            recommendation=(
                "Use only predefined strategy labels or update STRATEGY_OPTIONS."
                if len(invalid_strategy_df)
                else "No invalid strategy labels found."
            )
        )

    # Final summary sheet
    report["health_summary"] = pd.DataFrame(
        summary_records,
        columns=["check", "status", "issue_count", "recommendation"]
    )

    return report


def display_health_report(health_report):
    """
    Display the most important health report tables in Colab.
    """
    print("Health summary")
    display(health_report["health_summary"])

    print("\nOverview")
    display(health_report["overview"])

    for key, table in health_report.items():
        if key in ["health_summary", "overview"]:
            continue

        if isinstance(table, pd.DataFrame) and not table.empty:
            print(f"\n{key}")
            display(table)


def export_health_report(excel_path, health_report):
    """
    Export the corpus health report to an Excel file.
    """
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        # Put summary first
        if "health_summary" in health_report:
            health_report["health_summary"].to_excel(
                writer,
                index=False,
                sheet_name="health_summary"
            )

        if "overview" in health_report:
            health_report["overview"].to_excel(
                writer,
                index=False,
                sheet_name="overview"
            )

        for sheet_name, table in health_report.items():
            if sheet_name in ["health_summary", "overview"]:
                continue

            if isinstance(table, pd.DataFrame):
                # Excel sheet names must be <= 31 characters
                safe_sheet_name = sheet_name[:31]

                table.to_excel(
                    writer,
                    index=False,
                    sheet_name=safe_sheet_name
                )

    format_hanparal_workbook(excel_path)

    return excel_path